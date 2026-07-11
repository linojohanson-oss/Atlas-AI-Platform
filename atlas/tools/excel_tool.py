from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from atlas.tools.base_tool import BaseTool


class ExcelTool(BaseTool):
    """
    Herramienta para inspeccionar y analizar archivos Excel.

    Permite:
    - listar hojas;
    - seleccionar una hoja;
    - obtener dimensiones;
    - detectar encabezados;
    - mostrar una vista previa;
    - calcular estadísticas básicas de columnas numéricas.
    """

    name = "excel"
    description = (
        "Abre archivos Excel, inspecciona hojas, muestra datos "
        "y calcula estadísticas básicas."
    )
    version = "1.0.0"

    def execute(self, **kwargs: Any) -> Dict[str, Any]:
        """
        Ejecuta el análisis de un archivo Excel.
        """
        file_path_value = str(
            kwargs.get("file_path", "")
        ).strip()

        sheet_name_value = kwargs.get("sheet_name")
        preview_rows = int(
            kwargs.get("preview_rows", 5)
        )

        if not file_path_value:
            raise ValueError(
                "Debe indicar la ruta del archivo Excel."
            )

        if preview_rows < 1:
            raise ValueError(
                "preview_rows debe ser mayor que cero."
            )

        file_path = self._resolve_file_path(
            file_path_value
        )

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        try:
            sheet_names = workbook.sheetnames

            if not sheet_names:
                raise ValueError(
                    "El archivo Excel no contiene hojas."
                )

            selected_sheet_name = (
                str(sheet_name_value).strip()
                if sheet_name_value
                else sheet_names[0]
            )

            if selected_sheet_name not in sheet_names:
                available = ", ".join(sheet_names)

                raise ValueError(
                    f"La hoja '{selected_sheet_name}' no existe. "
                    f"Hojas disponibles: {available}"
                )

            worksheet = workbook[selected_sheet_name]

            rows = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

            if not rows:
                return self._empty_result(
                    file_path=file_path,
                    sheet_names=sheet_names,
                    selected_sheet=selected_sheet_name,
                )

            headers = self._build_headers(
                rows[0]
            )

            data_rows = rows[1:]

            preview = self._build_preview(
                headers=headers,
                rows=data_rows,
                limit=preview_rows,
            )

            numeric_summary = (
                self._calculate_numeric_summary(
                    headers=headers,
                    rows=data_rows,
                )
            )

            return {
                "tool": self.name,
                "status": "completed",
                "input": {
                    "file_path": str(file_path),
                    "sheet_name": selected_sheet_name,
                    "preview_rows": preview_rows,
                },
                "result": {
                    "file_name": file_path.name,
                    "file_path": str(file_path),
                    "sheet_names": sheet_names,
                    "selected_sheet": selected_sheet_name,
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                    "headers": headers,
                    "preview": preview,
                    "numeric_summary": numeric_summary,
                },
            }

        finally:
            workbook.close()

    @staticmethod
    def _resolve_file_path(
        file_path_value: str,
    ) -> Path:
        """
        Resuelve y valida la ruta del archivo.
        """
        file_path = Path(
            file_path_value
        ).expanduser()

        if not file_path.is_absolute():
            file_path = Path.cwd() / file_path

        file_path = file_path.resolve()

        if not file_path.exists():
            raise FileNotFoundError(
                f"El archivo no existe: {file_path}"
            )

        if not file_path.is_file():
            raise ValueError(
                f"La ruta no corresponde a un archivo: "
                f"{file_path}"
            )

        if file_path.suffix.lower() != ".xlsx":
            raise ValueError(
                "ExcelTool acepta únicamente archivos .xlsx."
            )

        return file_path

    @staticmethod
    def _build_headers(
        first_row: tuple,
    ) -> List[str]:
        """
        Construye nombres de columnas válidos y únicos.
        """
        headers: List[str] = []
        used_headers: Dict[str, int] = {}

        for index, value in enumerate(
            first_row,
            start=1,
        ):
            header = (
                str(value).strip()
                if value is not None
                else f"columna_{index}"
            )

            if not header:
                header = f"columna_{index}"

            if header in used_headers:
                used_headers[header] += 1
                header = (
                    f"{header}_{used_headers[header]}"
                )
            else:
                used_headers[header] = 1

            headers.append(header)

        return headers

    @staticmethod
    def _build_preview(
        headers: List[str],
        rows: List[tuple],
        limit: int,
    ) -> List[Dict[str, Any]]:
        """
        Convierte las primeras filas a diccionarios.
        """
        preview: List[Dict[str, Any]] = []

        for row in rows[:limit]:
            normalized_row = list(row)

            if len(normalized_row) < len(headers):
                normalized_row.extend(
                    [None]
                    * (
                        len(headers)
                        - len(normalized_row)
                    )
                )

            preview.append(
                {
                    header: normalized_row[index]
                    if index < len(normalized_row)
                    else None
                    for index, header in enumerate(
                        headers
                    )
                }
            )

        return preview

    @staticmethod
    def _calculate_numeric_summary(
        headers: List[str],
        rows: List[tuple],
    ) -> Dict[str, Dict[str, Any]]:
        """
        Calcula estadísticas básicas por columna numérica.
        """
        summary: Dict[str, Dict[str, Any]] = {}

        for column_index, header in enumerate(
            headers
        ):
            numeric_values: List[float] = []

            for row in rows:
                if column_index >= len(row):
                    continue

                value = row[column_index]

                if isinstance(value, bool):
                    continue

                if isinstance(value, (int, float)):
                    numeric_values.append(
                        float(value)
                    )

            if not numeric_values:
                continue

            summary[header] = {
                "count": len(numeric_values),
                "sum": sum(numeric_values),
                "mean": mean(numeric_values),
                "minimum": min(numeric_values),
                "maximum": max(numeric_values),
            }

        return summary

    def _empty_result(
        self,
        file_path: Path,
        sheet_names: List[str],
        selected_sheet: str,
    ) -> Dict[str, Any]:
        """
        Devuelve una respuesta válida para una hoja vacía.
        """
        return {
            "tool": self.name,
            "status": "completed",
            "input": {
                "file_path": str(file_path),
                "sheet_name": selected_sheet,
            },
            "result": {
                "file_name": file_path.name,
                "file_path": str(file_path),
                "sheet_names": sheet_names,
                "selected_sheet": selected_sheet,
                "row_count": 0,
                "column_count": 0,
                "headers": [],
                "preview": [],
                "numeric_summary": {},
            },
        }